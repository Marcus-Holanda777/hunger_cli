"""
   BUG: Removido do projeto
        self.service = FirefoxService(executable_path=GeckoDriverManager().install())
        self.driver = webdriver.Firefox(service=self.service, options=self.options)
        
        from selenium.webdriver.firefox.service import Service as FirefoxService
        from webdriver_manager.firefox import GeckoDriverManager
    
    * Selenium atualizado para a versao 4.14
"""
import mimetypes
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.remote.webelement import WebElement
from time import sleep
from random import randint
from dataclasses import dataclass
from datetime import datetime
from openpyxl import load_workbook
import pandas as pd
from os import makedirs

# interacao com o teclado e mouse
from selenium.webdriver import ActionChains, Keys

# esperar a tag aparecer
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

LIMIT_DAY = 31

MIME_TYPES = ";".join(mimetypes.types_map.values())
DT_REF = datetime.now()
PATH_DOWNLOAD = (
    Path() / "download_qu" /
    f"{DT_REF.day:02d}{DT_REF.month:02d}{DT_REF.year}"
)


@dataclass()
class Navegacao:
    by: By
    tag: str
    send_text: str = None
    btn: bool = False


class SpiderQu:
    def __init__(
        self, 
        url: str, 
        implicitly_wait: float,
        login: str, 
        password: str, 
        date_start: datetime, 
        date_end: datetime,
        invisible: bool = False
    ) -> None:

        self.url = url
        self.implicitly_wait = implicitly_wait
        self.user = login
        self.password = password
        self.date_start = date_start
        self.date_end = date_end
        self.invisible = invisible
        self.options = self.__options()
        self.driver = webdriver.Firefox(options=self.options)


    def __options(self):

        if not PATH_DOWNLOAD.is_dir():
            makedirs(PATH_DOWNLOAD)

        options = FirefoxOptions()
        options.set_preference("browser.download.folderList", 2)
        options.set_preference("browser.download.manager.showWhenStarting", False)
        options.set_preference("browser.download.dir", str(PATH_DOWNLOAD.absolute()))
        options.set_preference("browser.helperApps.neverAsk.saveToDisk", MIME_TYPES)
        if self.invisible:
            options.add_argument("-headless")

        return options
    

    def esperar_tag(
        self, 
        by: By, 
        tag: str, 
        timeout: float = 30.0, 
        all: bool = False
    ) -> list[WebElement] | WebElement:

        if not all:

            element = (
                WebDriverWait(driver=self.driver, timeout=timeout)
                .until(
                    EC.presence_of_element_located(
                        (by, tag)
                    )
                )
            )

        else:

            element = (
                WebDriverWait(driver=self.driver, timeout=timeout)
                .until(
                    EC.presence_of_all_elements_located(
                        (by, tag)
                    )
                )
            )


        return element
    

    def navegar(
        self, 
        tags: list[Navegacao]
    ) -> None:

        for tag in tags:
            action_tag = self.esperar_tag(tag.by, tag.tag)

            if not tag.btn:
                action_tag.send_keys(tag.send_text)
            else:
                action_tag.click()


    def delay(
        self, 
        start: int, 
        end: int
    ) -> None:

        espera =  randint(start, end)
        sleep(float(espera))
    

    def get_url(self) -> None:
        self.driver.get(self.url)
    

    def login(self) -> None:
        tags = [
            Navegacao(by=By.CSS_SELECTOR, tag="input[name='username']", send_text=self.user), 
            Navegacao(by=By.CSS_SELECTOR, tag="input[name='password']", send_text=self.password),
            Navegacao(by=By.CSS_SELECTOR, tag="button[name='login-btn']", btn=True)
        ]

        self.navegar(tags=tags)
        self.esperar_tag(By.XPATH, "//span[contains(text(), 'Close')]")
        
        escape = ActionChains(self.driver).send_keys(Keys.ESCAPE)
        escape.perform()

        self.delay(2, 4)

    
    def relatorio(self):
        tags = [
            Navegacao(by=By.PARTIAL_LINK_TEXT, tag="Reports", btn=True),
            Navegacao(by=By.XPATH, tag="//span[contains(text(), 'Summary By Date')]", btn=True)
        ]

        self.navegar(tags=tags)
        self.delay(2, 4)
    

    def date_from_to(self):
        # TODO: LOCALIZAR A LABEL FROM
        tags = ('From', 'To')

        for tag_label in tags:
            tbl_from = (
                self.esperar_tag(By.XPATH, f"//label[normalize-space(text())='{tag_label}']")
                 .find_element(By.XPATH, "..")
            )

            if tag_label == 'From':
                date_ref = self.date_start
            else:
                date_ref = self.date_end

            # TODO: Filtro YEAR
            elm_year = tbl_from.find_element(By.CSS_SELECTOR, f"option[value='{str(date_ref.year)}']").find_element(By.XPATH, "..")
            sel_year = Select(elm_year)
            sel_year.select_by_value(str(date_ref.year))

            # TODO: Filtro MONTHY
            elm_month = tbl_from.find_element(By.CSS_SELECTOR, f"option[value='{str(date_ref.month - 1)}']").find_element(By.XPATH, "..")
            sel_month = Select(elm_month)
            sel_month.select_by_value(str(date_ref.month - 1))

            # TODO: Filtro DAY
            elm_day = tbl_from.find_element(By.LINK_TEXT, str(date_ref.day))
            elm_day.click()

            self.delay(2, 4)
        
        btn_sucess = self.esperar_tag(By.XPATH, "//button[@class='button is-success']")
        btn_sucess.click()


    def store_list(self) -> list[tuple[WebElement, str]]:
        # TODO: Botao para mostra a lista
        btn_list = self.esperar_tag(By.XPATH, "//input[@value='store-list']")
        ActionChains(self.driver).click(btn_list).perform()

        # TODO: Botao que mostra a lista de lojas
        self.btn_danger().click()
        self.delay(2, 4) # 2, 4

        # TODO: Capturar lista de lojas
        # classe anterior //a[contains(@class,'dropdown-item')]
        # classe atual -- o-acp__item
        list_store = self.esperar_tag(
            By.XPATH, 
            "//div[contains(@class,'acp__item')]", 
            all=True
        )

        # TODO: Retornar uma lista de tuplas
        # com o elemento e o texto
        # caso precise utilizar
        # 10/10/2023    
        list_all = [
            *zip(
                list_store, 
                map(
                    lambda el: el.text.strip(),
                    list_store
                )
            )
        ]

        # TODO: Excluir loja da lista
        return list_all
        

    def espera_download(self):
        original = self.driver.current_window_handle

        while True:
            is_visible = (
                WebDriverWait(self.driver, 30.0, 1.0)
                 .until_not(
                   EC.presence_of_element_located((By.XPATH, "//div[@class='modal-background']"))
                 )
            )

            is_quias = len(self.driver.window_handles) == 1

            if is_visible and is_quias:
                break

            sleep(2.0)
        
        self.driver.switch_to.window(original)
    

    def btn_danger(self):
        return (
            self.esperar_tag(By.XPATH, "//input[@placeholder='Select Store']")
        )


    def btn_export(self):
        return (
             self.esperar_tag(
                By.XPATH, 
                "//button[@class='button is-light']//span[normalize-space(text())='Export']"
            )
        )
    

    def btn_download(self):
        return (
            self.esperar_tag(
                By.XPATH, 
                "//button[@class='button is-light']//span[normalize-space(text())='Download File']"
            )
        )
            

    def download(
        self, 
        all_link: list[tuple[WebElement, str]]
    ) -> None:

        for __, text in all_link:
            self.btn_danger().click()
            self.delay(2, 4)
                
            # NOTE: Alterado dia 10/10/2023
            # Nao e mais um link
            # self.esperar_tag(By.LINK_TEXT, link).click()
            # O ELEMENTO esta sumindo do DRIVER
            link_new = self.esperar_tag(
                By.XPATH, 
                f'//div[normalize-space(text())="{text}"]'
            )
            link_new.click()

            self.btn_export().click()
            self.btn_download().click()
            self.espera_download()

            # NOTE: Alterado dia 10/10/2023
            # antes By.XPATH, "//a[@class='tag is-delete']")
            # alterado 'o-inputit__item__close
            # self.esperar_tag(By.XPATH, "//a[@class='tag is-delete']").click()
            self.esperar_tag(
                By.XPATH, 
                "//span[contains(@class, 'o-inputit__item__close')]"
            ).click()


    def filtros(self):

        filtro = self.esperar_tag(By.CSS_SELECTOR, "strong[data-msgid='Show Filters']")
        filtro.click()

        selm = self.esperar_tag(By.TAG_NAME, "select")
        
        list_sel = Select(selm)
        list_sel.select_by_value('range')
        
        # TODO: Definir intervalo de datas, para filtrar relatorio
        intervalo = self.esperar_tag(By.CSS_SELECTOR, "button[class='button is-info']")
        intervalo.click()
        
        self.delay(2, 4)
        self.date_from_to()

        all_link = self.store_list()
        self.download(all_link=all_link)


    def tratamento_base(
        self, 
        arq: Path
    ) -> pd.DataFrame:

        def name_location(arq: Path) -> str:
            wb = load_workbook(arq, read_only=True)
            ws = wb.active

            name = (
                ws.cell(6, 1)
                 .value
                 .split(':')[1]
                 .split(',')[0]
            )
            wb.close()

            return name
        
        def transform(df: pd.DataFrame) -> pd.DataFrame:
            return (
                pd
                  .melt(df, id_vars=['Section', 'Metric'], var_name="Date", value_name='Total')
                  .assign(Date = lambda _df: _df['Date'].map(pd.to_datetime), Location = name_location(arq))
              )
        
        not_total = "total"
        not_employees = "employees"

        return (
            pd
              .read_excel(arq, skiprows=7)
              .assign(Section = lambda _df: _df['Section'].fillna(method='ffill'))
              .iloc[:, :-1]
              .query("Metric.str.lower() != @not_total")
              .query("Section.str.lower() != @not_employees")
              .pipe(transform)
        )

    
    def concat_path(self):
        return ( 
            pd.concat(
                [
                    self.tratamento_base(arq)
                    for arq in PATH_DOWNLOAD.glob("**/*.xlsx")
                    if arq.stat().st_size > 0
                ]
            )
        )


    def close(self) -> None:
        self.driver.quit()
    

    def run(self) -> None:
        
        self.get_url()
        self.login()
        self.relatorio()
        self.filtros()
        self.close()

        self.delay(4,6)

        df = self.concat_path()
        df.to_excel('REPORT_QU.xlsx', index=False)